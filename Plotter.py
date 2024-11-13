from datetime import datetime

import astropy.time
import astropy.units as u
import hapsira.frames
import keyboard
import numpy as np
import pandas as pd
from astropy.time import Time, TimeDelta
from hapsira.bodies import Earth
from hapsira.ephem import Ephem
from hapsira.maneuver import Maneuver
from hapsira.plotting import OrbitPlotter
from hapsira.twobody import Orbit
from hapsira.util import time_range
from matplotlib import pyplot as plt
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook import Workbook

from JPLQuery import JPLQuery
from Mechanics import Mechanics


class Plotter:
    def __init__(self):
        pass

    @staticmethod
    def __generate_orbit(ephemerides: dict, EPOCH: astropy.time.Time, r1: float, rb: float,
                         plane: hapsira.frames.Planes, true_anomaly: float):
        a = (r1 + rb) / 2.0
        ecc = (r1 - rb) / (rb + r1)

        return Orbit.from_classical(
            attractor=Earth,
            plane=plane,
            epoch=EPOCH,
            a=a * u.km,
            ecc=ecc * u.one,
            inc=ephemerides['inc'] * (np.pi / 180.0) * u.rad,
            raan=ephemerides['raan'] * (np.pi / 180.0) * u.rad,
            argp=ephemerides['argp'] * (np.pi / 180.0) * u.rad,
            nu=true_anomaly * u.rad
        )

    @staticmethod
    def plot_orbit(body, years, r1=10000.0 + Earth.R.value / 1000.0, show=False):
        # Get the closest approach data (assuming JPLQuery is compatible with Hapsira)
        EPOCH_NOW = Time("2024-11-2", scale='tdb')
        Delta = TimeDelta(years * u.year, scale='tdb')
        deltaV = []
        deltaT = []

        beg = EPOCH_NOW - Delta
        end = EPOCH_NOW + Delta

        # Extract date and convert it to a suitable format
        date_str = JPLQuery.get_closest_approach(body,
                                                 beg.strftime("%Y-%m-%d"),
                                                 end.strftime("%Y-%m-%d"))['date'] + ":00"

        date_obj = datetime.strptime(date_str, '%Y-%b-%d %H:%M:%S')
        iso_format = date_obj.strftime('%Y-%m-%d %H:%M:%S')

        # Set the epoch for the orbit
        EPOCH = Time(iso_format, scale='tdb')

        # Earth.plot(EPOCH)
        epochs = time_range(
            start=EPOCH - TimeDelta(5 * u.day),
            end=EPOCH + TimeDelta(5 * u.day)
        )

        # ephemerides query
        orbit_epochs = {
            'start': (EPOCH - TimeDelta(5 * u.day)).strftime('%Y-%m-%d 00:00:00'),
            'stop': (EPOCH + TimeDelta(5 * u.day)).strftime('%Y-%m-%d 00:00:00'),
            'step': '1h'
        }

        ephemerides = JPLQuery.get_ephemerides(body, "500@399", orbit_epochs, EPOCH)
        orbital = Ephem.from_horizons(body, epochs, attractor=Earth)

        rb = (ephemerides['qr'] * 149597870.691) + Earth.R.value / 1000.0  # Convert AU to km

        orb_1 = Plotter.__generate_orbit(ephemerides, EPOCH, r1, r1, orbital.plane, np.pi / 2)
        orb_2 = Plotter.__generate_orbit(ephemerides, EPOCH, r1, rb, orbital.plane, 0)

        orb_1v = Mechanics.vel(1000 * r1, 1000 * r1)
        orb_2v = Mechanics.vel(1000 * (r1 + rb) / 2, 1000 * r1)

        deltaT.append(Mechanics.elliptic_period(1000 * (r1 + rb) / 2, orb_1.ecc, np.pi / 2, 0))
        deltaV.append((orb_2v - orb_1v) / 1000 * (u.km / u.s))

        # set the orbit frame to be in plane with the earth and the asteroid
        # frame.set_orbit_frame(orb_2)

        # Orbit back
        v_initial = Mechanics.vel(1000 * (r1 + rb) / 2, 1000 * rb)
        v_final = Mechanics.vel(1000 * r1, 1000 * r1)

        deltaV.append(abs(v_final - v_initial) / 1000 * (u.km / u.s))
        dv_vector = Mechanics.vec_in_dir(deltaV[1].value, orb_2.r)

        # TODO: fix instantaneous impulses and use engine thrust curves instead
        man = Maneuver((0 * u.s, dv_vector * (u.km / u.s)))
        orb_3 = orb_2.apply_maneuver(man)

        print(orb_3.a.value, orb_3.ecc)
        deltaT.append(Mechanics.hyperbolic_period(orb_3.a.value * 1000, orb_3.ecc, 0, -np.pi / 2))

        # orb 4
        man = Maneuver.hohmann(orb_3, r1 * u.km)
        deltaV.append(man.get_total_cost())
        orb_4 = orb_3.apply_maneuver(man)

        if show:
            frame = OrbitPlotter()
            frame.set_attractor(Earth)
            frame.plot(orb_1, label="Orbit 1", color="black")
            frame.plot(orb_2, label="Orbit 2", color="orange")
            frame.plot(orb_3, label="Orbit 3", color="blue")
            frame.plot(orb_4, label="Orbit 4", color="green")
            frame.plot_ephem(orbital, epoch=EPOCH, label=body, color="red")

            plt.xlim(-rb * 0.75, rb * 1.5)
            plt.ylim(-rb, rb * 0.5)

            plt.title(f"Hohmann Transfers for {body}")
            plt.show()

        return deltaV, deltaT

    @staticmethod
    def plot_timed(body, start_time, end_time, r1=10 * 1000.0 + Earth.R.value / 1000.0, show=False):
        # Transfer Cost, Transfer Time
        deltaV = []
        deltaT = []

        query_epochs = {
            'start': start_time.strftime('%Y-%m-%d 00:00:00'),
            'stop': end_time.strftime('%Y-%m-%d 00:00:00'),
            'step': '1d'
        }

        # Get closest approach data
        data = JPLQuery.get_closest_approach(body,
                                             "500@399",
                                             query_epochs)
        date_str = data['date'] + " 00:00:00"
        date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        iso_format = date_obj.strftime('%Y-%m-%d %H:%M:%S')
        EPOCH = Time(iso_format, scale='tdb')

        # Calculate the Hohmann period. Get the Ephemerides of the body, and find the necessary orbital transfers.
        rb = (data['distance_au'] * 149597870.691) + Earth.R.value / 1000.0  # Convert AU to km
        deltaT.append(Mechanics.elliptic_period(1000 * (r1 + rb) / 2, 0, np.pi / 2, 0) * u.s)

        print(f"rb from JPL {rb}")
        # Plot the orbit of body so that it hits perigee at the end.
        Delta = TimeDelta(deltaT[0], scale='tdb')

        # Generate the orbital transfers.
        query_epochs = {
            'start': (EPOCH - Delta).strftime('%Y-%m-%d 00:00:00'),
            'stop': EPOCH.strftime('%Y-%m-%d 00:00:00'),
            'step': '1h'
        }
        orbital = Ephem.from_horizons(body, epochs=time_range(start=EPOCH - Delta, end=EPOCH), attractor=Earth)
        ephemerides = JPLQuery.get_ephemerides(body, "500@399", query_epochs, EPOCH)

        initial_orbit = Plotter.__generate_orbit(ephemerides, EPOCH, r1, r1, orbital.plane, np.pi / 2)
        rendezvous_orbit = Plotter.__generate_orbit(ephemerides, EPOCH, r1, rb, orbital.plane, 0)

        # Return Orbit Transfer requires the same cost
        v_initial = Mechanics.vel(1000 * r1, 1000 * r1)
        v_rendezvous = Mechanics.vel(1000 * (r1 + rb) / 2, 1000 * r1)
        v_return = Mechanics.vel(1000 * (r1 + rb) / 2, 1000 * rb)

        dv_vector = Mechanics.vec_in_dir(abs(v_initial - v_return) / 1000 * (u.km / u.s), rendezvous_orbit.r)
        return_orbit = rendezvous_orbit.apply_maneuver(Maneuver((0 * u.s, dv_vector * (u.km / u.s))))

        deltaV.append((v_rendezvous - v_initial) / 1000 * (u.km / u.s))
        deltaV.append(abs(v_return - v_initial) / 1000 * (u.km / u.s))
        deltaT.append(Mechanics.hyperbolic_period(return_orbit.a.value * 1000, return_orbit.ecc, 0, -np.pi / 2) * u.s)

        # Adjustment
        man = Maneuver.hohmann(return_orbit, r1 * u.km)
        deltaV.append(man.get_total_cost())
        deltaT.append(man.get_total_time())
        adjustment_orbit = return_orbit.apply_maneuver(man)

        if show:
            frame = OrbitPlotter()
            frame.set_attractor(Earth)

            frame.plot(initial_orbit, label="Initial Orbit", color="blue")
            frame.plot(rendezvous_orbit, label="Rendezvous Orbit", color="orange")
            frame.plot(return_orbit, label="Return Orbit", color="green")
            frame.plot(adjustment_orbit, label="Adjustment", color="yellow")
            frame.plot_ephem(orbital, epoch=EPOCH, label=body, color="red")

            plt.title(f"Hohmann Transfers for {body}")
            plt.show()

        return deltaV, deltaT


def to_excel(col_data, col_titles, path, title, x_axis_title, y_axis_title):
    # Create a DataFrame from the provided column data and titles
    data_frame_args = {col_titles[index]: col_data[index] for index in range(len(col_titles))}
    data = pd.DataFrame(data_frame_args)

    # Save the DataFrame to an Excel file
    data.to_excel(path, sheet_name=ast_name, index=False)

    # Open the existing workbook (instead of creating a new one)
    wb = Workbook()
    ws = wb.active

    ws.append(col_titles)
    for r in data.itertuples(index=False):
        ws.append(r)

    # Create a LineChart object
    chart = LineChart()

    # Set chart title and axis labels
    chart.title = title
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title

    # Define data reference for Y-axis (data range)
    data_ref = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)  # Assuming data starts from column 2

    # Define categories (X-axis labels)
    categories_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Assuming X-axis is in column 1

    # Add data and categories to the chart
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    # Add the chart to the worksheet at a specific position
    ws.add_chart(chart, "G5")

    # Save the workbook with the chart added
    wb.save(path)


if __name__ == "__main__":
    def get_sub_l(list_of_lists, index):
        return list(map(lambda values: values[index].value, list_of_lists))


    ast_name = "2024 PT5"
    vels = []
    times = []

    now = Time("2024-11-2", scale='tdb')
    time_delta = TimeDelta(10 * u.year, scale='tdb')

    beg = now - time_delta
    end = now + time_delta

    x = range(10, 101)
    for i in x:
        res = Plotter.plot_timed(ast_name, beg, end, r1=i * 1000.0 + Earth.R.value / 1000.0, show=True)
        print(res)
        vels.append(res[0])
        times.append(sum(res[1]).value / (3600 * 24))

        input("continue...")
        if keyboard.is_pressed('q'):
            break

    magnitudes = [sum(x).value for x in vels]

    to_excel([list(x), get_sub_l(vels, 0), get_sub_l(vels, 1), get_sub_l(vels, 2), magnitudes, times],
             ["x", "V1", "V2", "V3", "|V|, T"],
             'data.xlsx',
             'Cost vs Initial LEO Orbit',
             'Initial Orbital Distance (1e3 km)',
             'Total Cost (km/s)')

    # Create the figure and axis
    fig, ax1 = plt.subplots()

    # Plot the total cost on the primary y-axis (left)
    ax1.plot(list(x), magnitudes, 'b-', label='Total Cost')

    ax1.set_xlabel('Initial Orbital Radius (1000 km)')
    ax1.set_ylabel('Total Cost (km/s)', color='b')
    ax1.tick_params(axis='y', labelcolor='b')
    ax1.legend()

    # Create a secondary y-axis for time in days
    ax2 = ax1.twinx()

    ax2.plot(list(x), times, 'r-', label='Time of Flight')
    ax2.set_ylabel('Time of Flight (days)', color='r')
    ax2.tick_params(axis='y', labelcolor='r')
    ax2.legend()

    # Add title and show the plot
    plt.title('Total Cost vs Initial Orbital Radius with Time of Flight')

    plt.legend()
    fig.tight_layout()  # Adjust layout to prevent overlap
    plt.show()
